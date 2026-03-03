"""Excel export for attendance sheet."""
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

from .models import Class, Column, Row, Cell, MonthConfig, Holiday, Attendance


def build_attendance_sheet(class_id, month, year):
    """Build one sheet for given class and month. Returns openpyxl Workbook."""
    try:
        class_obj = Class.objects.get(pk=class_id)
    except Class.DoesNotExist:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = _month_name(month)

    try:
        month_cfg = MonthConfig.objects.get(class_ref=class_obj, month=month, year=year)
        num_days = month_cfg.num_days
    except MonthConfig.DoesNotExist:
        num_days = 31

    ws.merge_cells('A1:C1')
    ws['A1'] = class_obj.school_name or ''
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:C2')
    ws['A2'] = f"CLASS - {class_obj.class_label or class_obj.name}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:C3')
    ws['A3'] = f"CLASS INCHARGE - {class_obj.class_incharge or ''}"
    ws['A3'].alignment = Alignment(horizontal='center')

    columns = list(class_obj.columns.all().order_by('order', 'id'))
    rows = list(class_obj.rows.all().order_by('row_order', 'id'))

    col_idx = 1
    for col in columns:
        ws.cell(row=5, column=col_idx, value=col.heading)
        col_idx += 1
    for day in range(1, num_days + 1):
        ws.cell(row=5, column=col_idx, value=day)
        col_idx += 1
    ws.cell(row=5, column=col_idx, value="Total Present")
    col_idx += 1
    ws.cell(row=5, column=col_idx, value="Total Absent")

    holiday_days = set(
        h.day for h in Holiday.objects.filter(class_ref=class_obj, month=month, year=year)
    )
    holiday_reasons = {
        h.day: h.reason for h in Holiday.objects.filter(class_ref=class_obj, month=month, year=year)
    }
    absent_set = set(
        (a.row_id, a.day)
        for a in Attendance.objects.filter(row__class_ref=class_obj, month=month, year=year)
    )
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='00FFFFFF', bold=True)

    for r_idx, row_obj in enumerate(rows, start=6):
        col_idx = 1
        for col in columns:
            cell = Cell.objects.filter(row=row_obj, column=col).first()
            val = cell.value if cell else ''
            ws.cell(row=r_idx, column=col_idx, value=val)
            col_idx += 1

        present_count = 0
        absent_count = 0
        for day in range(1, num_days + 1):
            if day in holiday_days:
                reason = holiday_reasons.get(day, 'H')
                char_idx = (r_idx - 6) // 2
                if (r_idx - 6) % 2 == 0 and char_idx < len(reason):
                    val = reason[char_idx].upper()
                else:
                    val = ''
                c = ws.cell(row=r_idx, column=col_idx, value=val)
                if val:
                    c.fill = red_fill
                    c.font = white_font
            elif (row_obj.id, day) in absent_set:
                val = 'A'
                c = ws.cell(row=r_idx, column=col_idx, value=val)
                c.fill = red_fill
                c.font = white_font
                absent_count += 1
            else:
                val = 'P'
                ws.cell(row=r_idx, column=col_idx, value=val)
                present_count += 1
            col_idx += 1

        ws.cell(row=r_idx, column=col_idx, value=present_count)
        col_idx += 1
        ws.cell(row=r_idx, column=col_idx, value=absent_count)

    return wb


def _month_name(month):
    names = ['', 'JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
             'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']
    return names[month] if 1 <= month <= 12 else str(month)


def export_excel(class_id, month, year):
    wb = build_attendance_sheet(class_id, month, year)
    if wb is None:
        return None
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
